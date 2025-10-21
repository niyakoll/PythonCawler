import time
import json
import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook, Workbook
from pathlib import Path
from datetime import date
from datetime import datetime
############################## Create Excel ########################################
def createExcel(filename,sheet_name):
    combined_data = {
    "ID":[],
    "keyword":[],
    "text":[],
    "likeCount":[],
    "replyCount":[],
    "link":[],
    "datetime":[],
    "recordTime":[],
    "type":[]

    
    }
    df = pd.DataFrame(combined_data)

    with pd.ExcelWriter(filename, engine = "openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name,index=False)

############################## Write Data to Excel #################################
def appendRecord(sheet,id,keyword,text,likeCount,replyCount, link ,datetime,recordTime,type):
    # Append data
    row = [id,keyword,text,likeCount,replyCount,link,datetime,recordTime,type]
    sheet.append(row)

############################## Update Data to Excel #################################
def updatePostInfo(df,id,keyword,text,likeCount,replyCount, link ,datetime,recordTime,type):
    index = findPostByID(df=df,id=id)
    df.loc[index] = {"ID":id,"keyword":keyword,"text":text,"likeCount":likeCount,"replyCount":replyCount,"link":link ,"datetime":datetime,"recordTime":recordTime,"type":type}

def findPostByID(df,id):
    
    column_name = "ID"  # Replace with the actual column name
    search_string = id # Replace with the string you are looking for
    matching_all_rows = df.loc[(df[column_name] == search_string) & (df["type"] == "post") ]
    #matching_one_rows = matching_all_rows.iloc[0]
    #testResult = matching_one_rows["ID"]
    postIndex = matching_all_rows.index[0]
    return postIndex
############################## Read Data to Excel ####################################




############################## Save Excel ############################################
#Save Excel
def saveExcel(filename,sheet_name,df):
    with pd.ExcelWriter(path=filename, engine = "openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name,index=False)


if __name__ == "__main__":
    #basic set up for pandas excel, connect workbook,sheet and 
    #namming the file by day and month
    #today = date.today()
    #now = datetime.now()
    #current_month_full_name = now.strftime('%B')
    
    filename = f"Data.xlsx"
    sheet_name = f"Data"

    if os.path.exists(filename):
        workbook = load_workbook(filename)
    else:
        createExcel(filename=filename,sheet_name=sheet_name)
        workbook = load_workbook(filename)
    # Create or get sheet
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        # Add headers
        sheet.append(["ID","keyword", "text", "likeCount","replyCount","link","datetime","recordTime","type"])
    else:
        sheet = workbook[sheet_name]

    df = pd.read_excel(filename,sheet_name)

    
    #add one record
    appendRecord(sheet=sheet,
                id="3747507342763244724_68163591692",
                keyword="",
                text="阿bob今日食咗咩",
                replyCount=1000,
                likeCount=990,
                link="https://www.threads.com/@missmuthk/post/DQCFAEQE34R",
                datetime="10-14-2025 10:12:23 HKT",
                recordTime="10-14-2025 15:02:52 HKT",
                type="post")
    
    
    #update record with existing id
    updatePostInfo(df,
                    id="3747507342763244724_68163591692",
                    keyword="",
                    text="阿bob唱歌好好聽",
                    replyCount=1000,
                    likeCount=990,
                    link="https://www.threads.com/@missmuthk/post/DQCFAEQE34R",
                    datetime="10-14-2025 10:12:23 HKT",
                    recordTime="10-14-2025 15:02:52 HKT",
                    type="post")
    

    #Save Excel
    saveExcel(filename,sheet_name,df)


    #with open(currentPath+"/result/searchResult10.json","r",encoding="utf-8") as f:
    #data = json.loads(replies,ensure_ascii=False)
    
