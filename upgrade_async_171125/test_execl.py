
import time
import json
import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook, Workbook
from pathlib import Path
from datetime import date
from datetime import datetime
import random
import json
from nested_lookup import nested_lookup
from datetime import datetime
import pytz
import time
import os
import threads_main
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup  #网页解析，获取数据
import re  #正则表达式，进行文字匹配
import urllib.request, urllib.error  #制定URL，获取网页数据
#import xlwt  #进行excel操作
import requests
import json
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
import excel_data as ex
import result_text_cleaning as cln
############################## Create Excel ########################################
def createExcel(filename,sheet_name):
    try:
        combined_data = {
        "ID":[],
        "keyword":[],
        "text":[],
        "likeCount":[],
        "replyCount":[],
        "link":[],
        "datetime":[],
        "recordTime":[],
        "type":[]

        
        }
        df = pd.DataFrame(combined_data)

        with pd.ExcelWriter(filename, engine = "openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name,index=False)
    except Exception as e:
        print()


############################## Write Data to Excel #################################
def appendRecord(filename,sheet,id,keyword,text,likeCount,replyCount, link ,datetime,recordTime,type):
    try:
        if os.path.exists(filename):
            workbook = load_workbook(filename)
        else:
            createExcel(filename=filename,sheet_name=sheet_name)
            workbook = load_workbook(filename)
        # Create or get sheet
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            # Add headers
            sheet.append(["ID","keyword", "text", "likeCount","replyCount","link","datetime","recordTime","type"])
        else:
            sheet = workbook[sheet_name]
    except Exception as e:
        print(e)

    # Append data
    try:
        row = [id,keyword,text,likeCount,replyCount,link,datetime,recordTime,type]
        sheet.append(row)
        workbook.save(filename)
        print("success!")
        return True
    except:    
        return False
############################## Update Data to Excel #################################
def updatePostInfo(filename,sheet_name,id,keyword,text,likeCount,replyCount, link ,datetime,recordTime,type):
    try:
        if os.path.exists(filename):
            workbook = load_workbook(filename)
        else:
            createExcel(filename=filename,sheet_name=sheet_name)
            workbook = load_workbook(filename)
        # Create or get sheet
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            # Add headers
            sheet.append(["ID","keyword", "text", "likeCount","replyCount","link","datetime","recordTime","type"])
        else:
            sheet = workbook[sheet_name]
    except Exception as e:
        print(e)
    try:
        df = pd.read_excel(filename,sheet_name)
        index = findPostByID(df=df,id=id)
        if index != False:
            df.loc[index] = df.loc[index].astype(object)
            #df.loc[index] = None
            df.loc[index] = {"ID":id,"keyword":keyword,"text":text,"likeCount":likeCount,"replyCount":replyCount,"link":link ,"datetime":datetime,"recordTime":recordTime,"type":type}
            #Save Excel
            saveExcel(filename,sheet_name,df)
            print("success!")
            return True
        
        else:
            #print("no record!")
            return False
    except Exception as e:
        print(e)
def findPostByID(df,id):
    
    column_name = "ID"  # Replace with the actual column name
    search_string = id # Replace with the string you are looking for
    try:
        matching_all_rows = df.loc[(df[column_name] == search_string) & (df["type"] == "post") ]
        #matching_one_rows = matching_all_rows.iloc[0]
        #testResult = matching_one_rows["ID"]
        postIndex = matching_all_rows.index[0]
        print("success!")
        return postIndex    
    except Exception as e:
        print(e)
        print("no record!")
        return False
    
############################## Read Data to Excel ####################################
def findAttrById(filename,sheet_name,id):
    try:
        if os.path.exists(filename):
            workbook = load_workbook(filename)
        else:
            createExcel(filename=filename,sheet_name=sheet_name)
            workbook = load_workbook(filename)
        # Create or get sheet
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            # Add headers
            sheet.append(["ID","keyword", "text", "likeCount","replyCount","link","datetime","recordTime","type"])
        else:
            sheet = workbook[sheet_name]
    except Exception as e:
        print(e)
    try:
        attr = {
            "id":"",
            "keyword":"",
            "text":"",
            "likeCount":0,
            "replyCount":0,
            "link":"",
            "datetime":"",
            "recordTime":"",
            "type":""
        }
        df = pd.read_excel(filename,sheet_name)
        column_name = "ID"  # Replace with the actual column name
        search_string = id # Replace with the string you are looking for
        try:
            matching_all_rows = df.loc[(df[column_name] == search_string) & (df["type"] == "post") ]
            matching_one_rows = matching_all_rows.iloc[0]
            #testResult = matching_one_rows["ID"]
            #postIndex = matching_all_rows.index[0]
            attr["id"] = matching_one_rows["ID"]
            attr["keyword"] = matching_one_rows["keyword"]
            attr["text"] = matching_one_rows["text"]
            attr["likeCount"] = matching_one_rows["likeCount"].astype("int64")
            attr["replyCount"] = matching_one_rows["replyCount"].astype("int64")
            attr["link"] = matching_one_rows["link"]
            attr["datetime"] = matching_one_rows["datetime"]
            attr["recordTime"] = matching_one_rows["recordTime"]
            attr["type"] = matching_one_rows["type"]
            #print(attr)
            return attr    
        except Exception as e:
            print(e)
            print("no record!")
            return False


    except Exception as e:
        print(e)
    
    



############################## Save Excel ############################################
#Save Excel
def saveExcel(filename,sheet_name,df):
    with pd.ExcelWriter(path=filename, engine = "openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name,index=False)


if __name__ == "__main__":
    #basic set up for pandas excel, connect workbook,sheet and 
    #namming the file by day and month
    #today = date.today()
    #now = datetime.now()
    #current_month_full_name = now.strftime('%B')
    
    filename = f"testData.xlsx"
    sheetName = f"Data"

    """
    for i in range(40):
        like = random.randint(0,101)
        reply = random.randint(101,201)
        appendRecord(filename=filename,sheet=sheet_name,
                                        id=f"{i}",
                                        keyword="阿bob",
                                        text=f"{i}",
                                        replyCount=reply,
                                        likeCount=like,
                                        link=f"{i}",
                                        datetime=f"{i}",
                                        recordTime=f"{i}",
                                        type="post"
                     )
    """
    with open("C:/Users/Alex/ListeningTool/github/threads_template/result/inspectBug.json", 'r',encoding="utf-8") as file:
            result = json.load(file)
            if result != []:
                #print(len(posts))
                i = 0
                for post in result[:]:
                    postUrl = post["thread"]["url"]
                    postTimeStamp = post["thread"]["published_on"]
                    
                    postKeyword = post["thread"]["keyword"]
                    
                    postTitle = post["thread"]["text"]
                    
                    postId = post["thread"]["id"]
                    now = cln.timestampConvert(time.time())
                    attr = ex.findAttrById(filename=filename,sheet_name=sheetName,id=postId)
                    
                    if attr == False:
                        postTime = cln.timestampConvert(postTimeStamp)
                        postLikeCount = post["thread"]["like_count"]
                        postReplyCount = post["thread"]["direct_reply_count"]
                        #inputAiText += f"關鍵字:{postKeyword}\n帖文標題: {postTitle}\n發佈時間: {postTime}\n帖文連結: {postUrl}\n帖文讚好數: {postLikeCount}\n帖文留言數: {postReplyCount}\n"
                        try:
                            ex.appendRecord(
                                        filename=filename,
                                        sheet=sheetName,
                                        id=postId,
                                        keyword=postKeyword,
                                        text=postTitle,
                                        replyCount=postReplyCount,
                                        likeCount=postLikeCount,
                                        link=postUrl,
                                        datetime=postTime,
                                        recordTime=now,
                                        type="post"
                                        )
                        except Exception as e:
                            print(e)
                        #updatePostList += f"帖文標題: {postTitle}\n發佈時間: {postTime}\n帖文連結: {postUrl}\n帖文讚好數: {postLikeCount}\n帖文留言數: {postReplyCount}\n"
                        #print(f"Post Title: {postTitle}\n")
                        #print(f"Publish on: {postTime}\n")
                        #print(f"Post Link: {postUrl}\n")
                        #print(f"Post Like Count: {postLikeCount}\n")
                        #print(f"Post Reply Count: {postReplyCount}\n")
                        #replies = nested_lookup("replies",result)
                        reply = post["replies"]
                        #commentList = replies[i]
                        #for comment in commentList:
                        for comment in reply:
                            commentText = comment["text"]
                            commentTimeStamp = comment["published_on"]
                            commentTime = cln.timestampConvert(commentTimeStamp)
                            commentLikeCount = comment["like_count"]
                            commentUrl = comment["url"]
                            commentReplyCount = comment["direct_reply_count"]
                            #inputAiText += f"留言: {commentText}(發佈時間: {commentTime})\n留言讚好數: {commentLikeCount}\n留言回覆數: {commentReplyCount}\n"
                            #print(f"Comment Text: {commentText}(Publish on {commentTime})\n")
                            #print(f"Comment Like Count: {commentLikeCount}\n")
                            #print(f"Comment Reply Count: {commentReplyCount}\n")
                            try:
                                ex.appendRecord(
                                        filename=filename,
                                        sheet=sheetName,
                                        id=postId,
                                        keyword=postKeyword,
                                        text=commentText,
                                        replyCount=commentReplyCount,
                                        likeCount=commentLikeCount,
                                        link=commentUrl,
                                        datetime=commentTime,
                                        recordTime=now,
                                        type="reply"
                                        )
                            except Exception as e:
                                print(e)
                        
                    elif attr != False:
                        try:
                            if attr["likeCount"] == postLikeCount and attr["replyCount"] == postReplyCount:
                                print("This Post has no update!")
                            else:
                                print("add")
                                #inputAiText += f"關鍵字:{postKeyword}\n帖文標題: {postTitle}\n發佈時間: {postTime}\n帖文連結: {postUrl}\n帖文讚好數: {postLikeCount}\n帖文留言數: {postReplyCount}\n"
                        except Exception as e:
                            print(e)
                        finally:
                            ex.updatePostInfo(filename=filename,
                                            sheet_name=sheetName,
                                            id=postId,
                                            keyword=postKeyword,
                                            text=postTitle,
                                            replyCount=postReplyCount,
                                            likeCount=postLikeCount,
                                            link=postUrl,
                                            datetime=postTime,
                                            recordTime=now,
                                            type="post")
                        i+=1
    

    

    """
    for i in range(10):
        #add one record
        appendRecord(filename=filename,
                    sheet=sheet_name,
                    id="3747507342763244724_68163591699",
                    keyword="",
                    text="食咩好",
                    replyCount=1000,
                    likeCount=990,
                    link="https://www.threads.com/@missmuthk/post/DQCFAEQE34R",
                    datetime="10-14-2025 10:12:23 HKT",
                    recordTime="10-14-2025 15:02:52 HKT",
                    type="post")
        
        #update record with existing id
        updatePostInfo(filename=filename,
                    sheet_name=sheet_name,
                        id="3747507342763244724_68163591692",
                        keyword="",
                        text="阿bob",
                        replyCount=1000,
                        likeCount=990,
                        link="https://www.threads.com/@missmuthk/post/DQCFAEQE34R",
                        datetime="10-14-2025 10:12:23 HKT",
                        recordTime="10-14-2025 15:02:52 HKT",
                        type="post")
    
    """
    
    
    

    


    #with open(currentPath+"/result/searchResult10.json","r",encoding="utf-8") as f:
    #data = json.loads(replies,ensure_ascii=False)
    
