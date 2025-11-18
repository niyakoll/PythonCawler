
import time
import json
import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook, Workbook
from pathlib import Path
from datetime import date
from datetime import datetime
############################## Create Excel ########################################
def createExcel(filename,sheet_name):
    try:
        if os.path.exists(filename):
            workbook = load_workbook(filename)
        else:
            createExcel(filename=filename,sheet_name=sheet_name)
            workbook = load_workbook(filename)
        # Create or get sheet
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            # Add headers
            sheet.append(["ID","keyword", "text", "likeCount","replyCount","link","datetime","recordTime","type"])
        else:
            sheet = workbook[sheet_name]
    except Exception as e:
        print(e)
    try:
        combined_data = {
        "ID":[],
        "keyword":[],
        "text":[],
        "likeCount":[],
        "replyCount":[],
        "link":[],
        "datetime":[],
        "recordTime":[],
        "type":[]

        
        }
        df = pd.DataFrame(combined_data)

        with pd.ExcelWriter(filename, engine = "openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name,index=False)
    except Exception as e:
        print()


############################## Write Data to Excel #################################
def appendRecord(filename,sheet_name,id,keyword,text,likeCount,replyCount, link ,datetime,recordTime,type):
    try:
        if os.path.exists(filename):
            workbook = load_workbook(filename)
        else:
            createExcel(filename=filename,sheet_name=sheet_name)
            workbook = load_workbook(filename)
        # Create or get sheet
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            # Add headers
            sheet.append(["ID","keyword", "text", "likeCount","replyCount","link","datetime","recordTime","type"])
        else:
            sheet = workbook[sheet_name]
    except Exception as e:
        print(e)

    # Append data
    try:
        row = [id,keyword,text,likeCount,replyCount,link,datetime,recordTime,type]
        sheet.append(row)
        workbook.save(filename)
        print("success!")
        return True
    except:    
        return False
############################## Update Data to Excel #################################
def updatePostInfo(filename,sheet_name,id,keyword,text,likeCount,replyCount, link ,datetime,recordTime,type):
    try:
        if os.path.exists(filename):
            workbook = load_workbook(filename)
        else:
            createExcel(filename=filename,sheet_name=sheet_name)
            workbook = load_workbook(filename)
        # Create or get sheet
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            # Add headers
            sheet.append(["ID","keyword", "text", "likeCount","replyCount","link","datetime","recordTime","type"])
        else:
            sheet = workbook[sheet_name]
    except Exception as e:
        print(e)
    try:
        df = pd.read_excel(filename,sheet_name)
        index = findPostByID(df=df,id=id)
        if index != False:
            df.loc[index] = df.loc[index].astype(object)
            #df.loc[index] = None
            df.loc[index] = {"ID":id,"keyword":keyword,"text":text,"likeCount":likeCount,"replyCount":replyCount,"link":link ,"datetime":datetime,"recordTime":recordTime,"type":type}
            #Save Excel
            saveExcel(filename,sheet_name,df)
            print("success!")
            return True
        
        else:
            #print("no record!")
            return False
    except Exception as e:
        print(e)
def findPostByID(df,id):
    
    column_name = "ID"  # Replace with the actual column name
    search_string = id # Replace with the string you are looking for
    try:
        matching_all_rows = df.loc[(df[column_name] == search_string) & (df["type"] == "post") ]
        #matching_one_rows = matching_all_rows.iloc[0]
        #testResult = matching_one_rows["ID"]
        postIndex = matching_all_rows.index[0]
        print("success!")
        return postIndex    
    except Exception as e:
        print(e)
        print("no record!")
        return False
    
############################## Read Data to Excel ####################################
def findAttrById(filename,sheet_name,id):
    try:
        if os.path.exists(filename):
            workbook = load_workbook(filename)
        else:
            createExcel(filename=filename,sheet_name=sheet_name)
            workbook = load_workbook(filename)
        # Create or get sheet
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            # Add headers
            sheet.append(["ID","keyword", "text", "likeCount","replyCount","link","datetime","recordTime","type"])
        else:
            sheet = workbook[sheet_name]
    except Exception as e:
        print(e)
    try:
        attr = {
            "id":"",
            "keyword":"",
            "text":"",
            "likeCount":0,
            "replyCount":0,
            "link":"",
            "datetime":"",
            "recordTime":"",
            "type":""
        }
        df = pd.read_excel(filename,sheet_name)
        column_name = "ID"  # Replace with the actual column name
        search_string = id # Replace with the string you are looking for
        try:
            matching_all_rows = df.loc[(df[column_name] == search_string) & (df["type"] == "post") ]
            matching_one_rows = matching_all_rows.iloc[0]
            #testResult = matching_one_rows["ID"]
            #postIndex = matching_all_rows.index[0]
            attr["id"] = matching_one_rows["ID"]
            attr["keyword"] = matching_one_rows["keyword"]
            attr["text"] = matching_one_rows["text"]
            attr["likeCount"] = matching_one_rows["likeCount"].astype("int64")
            attr["replyCount"] = matching_one_rows["replyCount"].astype("int64")
            attr["link"] = matching_one_rows["link"]
            attr["datetime"] = matching_one_rows["datetime"]
            attr["recordTime"] = matching_one_rows["recordTime"]
            attr["type"] = matching_one_rows["type"]
            #print(attr)
            return attr    
        except Exception as e:
            print(e)
            print("no record!")
            return False


    except Exception as e:
        print(e)
    
    



############################## Save Excel ############################################
#Save Excel
def saveExcel(filename,sheet_name,df):
    with pd.ExcelWriter(path=filename, engine = "openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name,index=False)


if __name__ == "__main__":
    #basic set up for pandas excel, connect workbook,sheet and 
    #namming the file by day and month
    #today = date.today()
    #now = datetime.now()
    #current_month_full_name = now.strftime('%B')
    
    filename = f"Data.xlsx"
    sheet_name = f"Data"
    attr = findAttrById(filename=filename,sheet_name=sheet_name,id="3747507342763244724_681635916")
    print(attr) 
    """
    for i in range(10):
        #add one record
        appendRecord(filename=filename,
                    sheet=sheet_name,
                    id="3747507342763244724_68163591699",
                    keyword="",
                    text="食咩好",
                    replyCount=1000,
                    likeCount=990,
                    link="https://www.threads.com/@missmuthk/post/DQCFAEQE34R",
                    datetime="10-14-2025 10:12:23 HKT",
                    recordTime="10-14-2025 15:02:52 HKT",
                    type="post")
        
        #update record with existing id
        updatePostInfo(filename=filename,
                    sheet_name=sheet_name,
                        id="3747507342763244724_68163591692",
                        keyword="",
                        text="阿bob",
                        replyCount=1000,
                        likeCount=990,
                        link="https://www.threads.com/@missmuthk/post/DQCFAEQE34R",
                        datetime="10-14-2025 10:12:23 HKT",
                        recordTime="10-14-2025 15:02:52 HKT",
                        type="post")
    
    """
    
    
    

    


    #with open(currentPath+"/result/searchResult10.json","r",encoding="utf-8") as f:
    #data = json.loads(replies,ensure_ascii=False)
    
